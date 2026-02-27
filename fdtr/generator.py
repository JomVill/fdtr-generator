"""
FDTR Excel Generator
Produces a Faculty Daily Time Record .xlsx matching the MSU-IIT template exactly.
"""

import calendar
import io
from datetime import date, time
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

INSTITUTION_LINES = [
    "Republic of the Philippines",
    "Mindanao State University",
    "ILIGAN INSTITUTE OF TECHNOLOGY",
    "Iligan City",
]

MONTH_ABBR = {
    1: "Jan",  2: "Feb",  3: "Mar",  4: "Apr",
    5: "May",  6: "Jun",  7: "Jul",  8: "Aug",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

MONTH_FULL = {
    1: "January",   2: "February",  3: "March",    4: "April",
    5: "May",       6: "June",      7: "July",      8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}

WEEKDAY_NAMES = [
    "monday", "tuesday", "wednesday", "thursday",
    "friday", "saturday", "sunday",
]

# (time_in_col, time_out_col, hrs_col) — 1-indexed
CATEGORY_COLS = {
    "class":              (2,  3,  4),   # B, C, D
    "consultation":       (5,  6,  7),   # E, F, G
    "related_activities": (8,  9,  10),  # H, I, J
    "others":             (11, 12, 13),  # K, L, M
}

TOTAL_COL = 14   # N
DAY_COL   = 1    # A

# Row layout (matches Feb 2026 template)
COL_HEADER_ROW  = 10   # first column-header block (rows 10-13)
DATA_START_ROW  = 14   # day 1
SECTION2_HEADER = 63   # second column-header block (rows 63-66)
SECTION2_DATA   = 67   # day 16


# ---------------------------------------------------------------------------
# Border helpers  (all borders set BEFORE merge — slave cells are read-only)
# ---------------------------------------------------------------------------

_T = Side(style="thin")
_X = Side()   # no border

def _B(l=False, r=False, t=False, b=False) -> Border:
    """Build a Border from boolean flags."""
    return Border(
        left=_T   if l else _X,
        right=_T  if r else _X,
        top=_T    if t else _X,
        bottom=_T if b else _X,
    )

FULL = _B(l=True, r=True, t=True, b=True)   # ████
LR   = _B(l=True, r=True)                    # █ █
LRB  = _B(l=True, r=True,          b=True)   # █_█


def _frame_3rows(ws, col: int, r0: int) -> None:
    """
    Apply the 3-row "frame" border pattern to a single column (A or N):
      row r0    : full border
      row r0+1  : left + right only
      row r0+2  : left + right + bottom
    Must be called BEFORE merging.
    """
    ws.cell(r0,   col).border = FULL
    ws.cell(r0+1, col).border = LR
    ws.cell(r0+2, col).border = LRB


def _outline_merge_3rows(ws, r0: int, c0: int, c1: int) -> None:
    """
    Apply the 3-row outline border for a wide merged label (B–M).
    Row r0   : leftmost=full, interior=top-only, rightmost=right+top
    Row r0+1 : leftmost=left, rightmost=right
    Row r0+2 : leftmost=left+bottom, interior=bottom-only, rightmost=right+bottom
    Must be called BEFORE merging.
    """
    # Top row
    ws.cell(r0, c0).border = FULL
    for c in range(c0 + 1, c1):
        ws.cell(r0, c).border = _B(t=True)
    ws.cell(r0, c1).border = _B(r=True, t=True)
    # Middle row
    ws.cell(r0+1, c0).border = _B(l=True)
    ws.cell(r0+1, c1).border = _B(r=True)
    # Bottom row
    ws.cell(r0+2, c0).border = _B(l=True, b=True)
    for c in range(c0 + 1, c1):
        ws.cell(r0+2, c).border = _B(b=True)
    ws.cell(r0+2, c1).border = _B(r=True, b=True)


def _outline_merge(ws, r0: int, r1: int, c0: int, c1: int) -> None:
    """
    Generic outline border for any merged range. Set BEFORE merging.
    Only the four edges get borders (top row, bottom row, left col, right col).
    """
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            ws.cell(r, c).border = _B(
                l=(c == c0),
                r=(c == c1),
                t=(r == r0),
                b=(r == r1),
            )


# ---------------------------------------------------------------------------
# Alignment / Font constants  (name=None → system default = Calibri)
# ---------------------------------------------------------------------------

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _font(size=9, bold=False, italic=False) -> Font:
    return Font(name="Times New Roman", size=size, bold=bold, italic=italic)


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------

def _set(ws, row: int, col: int, value=None,
         font: Font = None, align: Alignment = None,
         border: Border = None, fmt: str = None):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    if fmt:
        cell.number_format = fmt
    return cell


def _merge(ws, r0: int, r1: int, c0: int, c1: int,
           value=None, font: Font = None, align: Alignment = None, fmt: str = None):
    """Merge cells and set top-left. Borders MUST be applied before calling."""
    ws.merge_cells(start_row=r0, end_row=r1, start_column=c0, end_column=c1)
    cell = ws.cell(row=r0, column=c0)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if fmt:
        cell.number_format = fmt
    return cell


# ---------------------------------------------------------------------------
# Time utilities
# ---------------------------------------------------------------------------

def _parse_time(t) -> Optional[time]:
    if t is None:
        return None
    if isinstance(t, time):
        return t
    if isinstance(t, str):
        try:
            h, m = t.strip().split(":")
            return time(int(h), int(m))
        except Exception:
            return None
    return None


def _hours(t_in, t_out) -> float:
    if t_in is None or t_out is None:
        return 0.0
    mins = (t_out.hour * 60 + t_out.minute) - (t_in.hour * 60 + t_in.minute)
    h = round(mins / 60, 2)
    return int(h) if h == int(h) else h


# ---------------------------------------------------------------------------
# Row address
# ---------------------------------------------------------------------------

def _day_row(day: int) -> int:
    """Return the first of the 3 rows for the given 1-indexed day."""
    if day <= 15:
        return DATA_START_ROW + (day - 1) * 3
    else:
        return SECTION2_DATA + (day - 16) * 3


# ---------------------------------------------------------------------------
# Header writers
# ---------------------------------------------------------------------------

def _merge_no_border(ws, r: int, c0: int, c1: int,
                     value=None, font: Font = None, align: Alignment = None):
    """Merge cells without any border (used for institution header rows 1-8)."""
    ws.merge_cells(start_row=r, end_row=r, start_column=c0, end_column=c1)
    cell = ws.cell(row=r, column=c0)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    return cell


def _write_institution_header(ws, month: int, year: int,
                               faculty_name: str, department: str):
    """
    Rows 1-9: institution name, title, month, name/dept.
    Rows 1-8: NO borders (plain text, merged).
    Row 9: borders start here (Name | Department).
    """
    # Rows 1-4: institution lines — no borders
    for i, line in enumerate(INSTITUTION_LINES, start=1):
        _merge_no_border(ws, i, 1, 14, value=line,
                         font=_font(10, bold=(i in (2, 3))), align=CENTER)

    # Row 5 blank — no border
    _merge_no_border(ws, 5, 1, 14)

    # Row 6: title — no border
    _merge_no_border(ws, 6, 1, 14,
                     value="FACULTY DAILY TIME RECORD",
                     font=_font(13, bold=True), align=CENTER)

    # Row 7: month — no border
    _merge_no_border(ws, 7, 1, 14,
                     value=f"For the month of {MONTH_FULL[month]} {year}",
                     font=_font(10), align=CENTER)

    # Row 8 blank — no border
    _merge_no_border(ws, 8, 1, 14)

    # Row 9: Name (A-H) | Dept (I-N) ← BORDERS START HERE
    _outline_merge(ws, 9, 9, 1, 8)
    _merge(ws, 9, 9, 1, 8,
           value=f"NAME: {faculty_name}",
           font=_font(9, bold=True), align=LEFT)
    _outline_merge(ws, 9, 9, 9, 14)
    _merge(ws, 9, 9, 9, 14,
           value=f"DEPARTMENT: {department}",
           font=_font(9, bold=True), align=LEFT)


def _write_col_headers(ws, start_row: int):
    """
    Write a 4-row column-header block (rows r … r+3).

    Row r  : DAY(A, ×4) | CLASS(B-D) | CONSULT(E-G) | RELATED(H-J, ×2) | OTHERS(K-M, ×2) | Total(N, ×4)
    Row r+1: (cont.)    | (cont.)    | (cont.)      | (merged above)    | (merged above)   |
    Row r+2: Time In | Time Out | Hrs  × 4  | Hrs
    Row r+3: In | Out |   × 4
    """
    r = start_row
    bf = _font(9, bold=True)
    sf = _font(8)

    # ── Row r ────────────────────────────────────────────────────────────
    _outline_merge(ws, r, r+3, 1, 1)
    _merge(ws, r, r+3, 1, 1, value="DAY", font=bf, align=CENTER)

    _outline_merge(ws, r, r, 2, 4)
    _merge(ws, r, r, 2, 4, value="CLASS", font=bf, align=CENTER)

    _outline_merge(ws, r, r, 5, 7)
    _merge(ws, r, r, 5, 7, value="CONSULTATION", font=bf, align=CENTER)

    _outline_merge(ws, r, r+1, 8, 10)
    _merge(ws, r, r+1, 8, 10, value="RELATED\nACTIVITIES", font=bf, align=CENTER)

    _outline_merge(ws, r, r+1, 11, 13)
    _merge(ws, r, r+1, 11, 13, value="OTHERS\n(Adm., R&E)", font=bf, align=CENTER)

    _outline_merge(ws, r, r+3, 14, 14)
    _merge(ws, r, r+3, 14, 14, value="Total\nHours", font=bf, align=CENTER)

    # ── Row r+1 ───────────────────────────────────────────────────────────
    _outline_merge(ws, r+1, r+1, 2, 4)
    _merge(ws, r+1, r+1, 2, 4, font=bf, align=CENTER)

    _outline_merge(ws, r+1, r+1, 5, 7)
    _merge(ws, r+1, r+1, 5, 7, font=bf, align=CENTER)

    # ── Row r+2 : Time In | Time Out | Hrs (per section) ─────────────────
    for c_start in (2, 5, 8, 11):
        _set(ws, r+2, c_start,   "Time In",  font=sf, align=CENTER, border=FULL)
        _set(ws, r+2, c_start+1, "Time Out", font=sf, align=CENTER, border=FULL)
        _set(ws, r+2, c_start+2, "Hrs",      font=sf, align=CENTER, border=FULL)

    # ── Row r+3 : In | Out (per section) ─────────────────────────────────
    for c_start in (2, 5, 8, 11):
        _set(ws, r+3, c_start,   "In",  font=sf, align=CENTER, border=FULL)
        _set(ws, r+3, c_start+1, "Out", font=sf, align=CENTER, border=FULL)
        _set(ws, r+3, c_start+2, None,  font=sf, align=CENTER, border=FULL)


# ---------------------------------------------------------------------------
# Day writers
# ---------------------------------------------------------------------------

def _write_special_day(ws, base_row: int, day_num: int,
                        label: str, total_value):
    """
    Write a weekend / holiday / leave / travel row.
    Structure:
      A  (col 1)  : 3-row merge — day number
      B–M (2-13)  : 3-row merge — label text
      N  (col 14) : 3-row merge — total (None for W/H, 0 for leave/travel)
    """
    r0 = base_row

    # ── A : day number ────────────────────────────────────────────────────
    _frame_3rows(ws, 1, r0)
    _merge(ws, r0, r0+2, 1, 1,
           value=day_num, font=_font(9, bold=True), align=CENTER)

    # ── B–M : label ───────────────────────────────────────────────────────
    _outline_merge_3rows(ws, r0, 2, 13)
    _merge(ws, r0, r0+2, 2, 13,
           value=label,
           font=_font(9, bold=True, italic=True),
           align=CENTER)

    # ── N : total ─────────────────────────────────────────────────────────
    _frame_3rows(ws, 14, r0)
    _merge(ws, r0, r0+2, 14, 14,
           value=total_value,
           font=_font(9, bold=True), align=CENTER)


def _write_regular_day(ws, base_row: int, day_num: int,
                        slots: list, force_category: str = None):
    """
    Write a regular workday.

    slots : list of dicts with keys: time_in, time_out, category, label
    force_category : if set (e.g. "related_activities"), override every
                     slot's category so all time goes into that column group.
    """
    r0 = base_row

    # ── A : day number (3-row merge) ──────────────────────────────────────
    _frame_3rows(ws, 1, r0)
    _merge(ws, r0, r0+2, 1, 1,
           value=day_num, font=_font(9, bold=True), align=CENTER)

    # ── B–M : individual cells — full thin border on ALL 3 sub-rows ───────
    for row_offset in range(3):
        r = r0 + row_offset
        for col in range(2, 14):
            ws.cell(r, col).border = FULL

    # ── Organise slots by category ────────────────────────────────────────
    cat_data: dict[str, list] = {k: [] for k in CATEGORY_COLS}
    total_hours = 0.0

    for slot in slots:
        cat = force_category if force_category else slot.get("category", "others")
        if cat not in cat_data:
            cat = "others"
        t_in  = _parse_time(slot.get("time_in"))
        t_out = _parse_time(slot.get("time_out"))
        hrs   = _hours(t_in, t_out)
        total_hours += hrs
        cat_data[cat].append((t_in, t_out, hrs))

    # Fill up to 2 entries per category into the 2 data sub-rows
    TIME_FMT = "h:MM AM/PM"
    for cat, entries in cat_data.items():
        cols = CATEGORY_COLS[cat]
        for sub_i, (t_in, t_out, hrs) in enumerate(entries[:2]):
            r = r0 + sub_i
            if t_in:
                _set(ws, r, cols[0], t_in,  font=_font(9), align=CENTER, fmt=TIME_FMT)
            if t_out:
                _set(ws, r, cols[1], t_out, font=_font(9), align=CENTER, fmt=TIME_FMT)
            if hrs:
                _set(ws, r, cols[2], hrs,   font=_font(9), align=CENTER)

    # ── N : total hours (3-row merge) ─────────────────────────────────────
    total_val = (int(total_hours) if total_hours == int(total_hours)
                 else round(total_hours, 2)) if total_hours > 0 else None

    _frame_3rows(ws, 14, r0)
    _merge(ws, r0, r0+2, 14, 14,
           value=total_val, font=_font(9, bold=True), align=CENTER)


# ---------------------------------------------------------------------------
# Footer
# ---------------------------------------------------------------------------

def _write_footer(ws, faculty_name: str, designation: str,
                  dept_head: str, month: int, year: int):
    r = SECTION2_DATA + 15 * 3   # row after last possible day (day 31)

    cert1 = ("This certifies upon my honor that the foregoing is a record "
             "for services I rendered to MSU-Iligan Institute")
    cert2 = f"of Technology during the month of {MONTH_FULL[month]} {year}."

    _merge(ws, r,   r,   1, 14, value=cert1, font=_font(8), align=LEFT)
    _merge(ws, r+1, r+1, 1, 14, value=cert2, font=_font(8), align=LEFT)

    # "Certified Correct:" (right side)
    _merge(ws, r+2, r+2, 10, 13,
           value="Certified Correct:", font=_font(8, bold=True), align=LEFT)

    sig = r + 5   # signature row

    # Faculty block
    _merge(ws, sig,   sig,   2, 6, value=faculty_name,
           font=_font(9, bold=True), align=CENTER)
    _merge(ws, sig+1, sig+1, 2, 6,
           value="(Signature over printed name)",
           font=_font(8), align=CENTER)
    _merge(ws, sig+2, sig+2, 2, 6, value=designation,
           font=_font(8), align=CENTER)
    _merge(ws, sig+3, sig+3, 2, 6, value="(Designation)",
           font=_font(8), align=CENTER)

    # Dept-head block
    _merge(ws, sig,   sig,   9, 13, value=dept_head,
           font=_font(9, bold=True), align=CENTER)
    _merge(ws, sig+1, sig+1, 9, 13,
           value="           (Head of Dept./Unit)",
           font=_font(8), align=CENTER)


# ---------------------------------------------------------------------------
# Column widths & row heights
# ---------------------------------------------------------------------------

def _set_column_widths(ws):
    ws.column_dimensions["A"].width = 5.0
    for col_letter in ("B", "C", "E", "F", "H", "I", "K", "L"):
        ws.column_dimensions[col_letter].width = 9.0
    for col_letter in ("D", "G", "J", "M"):
        ws.column_dimensions[col_letter].width = 5.0
    ws.column_dimensions["N"].width = 6.0


def _set_row_heights(ws, days_in_month: int):
    # Institution header rows
    for r in range(1, 10):
        ws.row_dimensions[r].height = 14.0
    ws.row_dimensions[6].height = 18.0   # FDTR title

    # Column-header blocks (10-13 and 63-66)
    for block in (COL_HEADER_ROW, SECTION2_HEADER):
        for offset in range(4):
            ws.row_dimensions[block + offset].height = 13.0

    # Day rows — all 3 sub-rows same height
    for day in range(1, days_in_month + 1):
        r0 = _day_row(day)
        ws.row_dimensions[r0].height   = 14.0
        ws.row_dimensions[r0+1].height = 14.0
        ws.row_dimensions[r0+2].height = 14.0


# ---------------------------------------------------------------------------
# Preview data generator (for HTML preview page)
# ---------------------------------------------------------------------------

def _fmt12(t: Optional[time]) -> str:
    """Format a time object as 12-hr string, e.g. time(8,30) → '8:30 AM'."""
    if t is None:
        return ""
    h    = t.hour
    m    = t.minute
    ampm = "AM" if h < 12 else "PM"
    h12  = h % 12 or 12
    return f"{h12}:{m:02d} {ampm}"


def _build_preview_regular(day: int, day_date, slots: list,
                            force_category: str = None) -> dict:
    """Build a regular-day dict for HTML preview rendering."""
    cat_data: dict = {k: [] for k in CATEGORY_COLS}
    total_hours = 0.0

    for slot in slots:
        cat = force_category if force_category else slot.get("category", "others")
        if cat not in cat_data:
            cat = "others"
        t_in  = _parse_time(slot.get("time_in"))
        t_out = _parse_time(slot.get("time_out"))
        hrs   = _hours(t_in, t_out)
        total_hours += hrs
        cat_data[cat].append({
            "in":  _fmt12(t_in),
            "out": _fmt12(t_out),
            "hrs": int(hrs) if hrs == int(hrs) else round(hrs, 2) if hrs else "",
        })

    total_val = ""
    if total_hours > 0:
        total_val = int(total_hours) if total_hours == int(total_hours) else round(total_hours, 2)

    day_class = "related" if force_category == "related_activities" else "regular"
    return {
        "day":       day,
        "date":      day_date,
        "type":      "regular",
        "day_class": day_class,
        "cat_data":  cat_data,
        "cats":      list(CATEGORY_COLS.keys()),
        "total":     total_val,
    }


def generate_preview_data(
    faculty_name: str,
    designation: str,
    department: str,
    dept_head: str,
    month: int,
    year: int,
    weekly_schedule: dict,
    special_days: dict,
) -> dict:
    """
    Return structured data for the HTML preview template.
    Same parameters as generate_fdtr().
    """
    days_in_month = calendar.monthrange(year, month)[1]
    rows = []

    for day in range(1, days_in_month + 1):
        day_date = date(year, month, day)
        date_str = day_date.strftime("%Y-%m-%d")
        weekday  = day_date.weekday()   # 0=Mon … 6=Sun

        if date_str in special_days:
            entry    = special_days[date_str]
            day_type = entry["type"]
            label    = entry.get("label", "")

            if day_type == "related_activities":
                custom_in  = entry.get("time_in",  "").strip()
                custom_out = entry.get("time_out", "").strip()
                if custom_in and custom_out:
                    slots = [{"time_in": custom_in, "time_out": custom_out,
                              "category": "related_activities", "label": ""}]
                else:
                    weekday_name = WEEKDAY_NAMES[weekday]
                    slots = weekly_schedule.get(weekday_name, [])
                rows.append(_build_preview_regular(day, day_date, slots,
                                                    force_category="related_activities"))
            elif day_type in ("leave", "travel"):
                rows.append({"day": day, "date": day_date, "type": "special",
                             "label": label, "total": 0, "day_class": day_type})
            else:  # holiday
                rows.append({"day": day, "date": day_date, "type": "special",
                             "label": label, "total": None, "day_class": "holiday"})

        elif weekday == 5:
            rows.append({"day": day, "date": day_date, "type": "special",
                         "label": "SATURDAY", "total": None, "day_class": "weekend"})
        elif weekday == 6:
            rows.append({"day": day, "date": day_date, "type": "special",
                         "label": "SUNDAY", "total": None, "day_class": "weekend"})
        else:
            weekday_name = WEEKDAY_NAMES[weekday]
            slots = weekly_schedule.get(weekday_name, [])
            rows.append(_build_preview_regular(day, day_date, slots))

    return {
        "faculty_name": faculty_name,
        "designation":  designation,
        "department":   department,
        "dept_head":    dept_head,
        "month":        month,
        "month_name":   MONTH_FULL[month],
        "year":         year,
        "days":         rows,
    }


# ---------------------------------------------------------------------------
# Main public function
# ---------------------------------------------------------------------------

def generate_fdtr(
    faculty_name: str,
    designation: str,
    department: str,
    dept_head: str,
    month: int,
    year: int,
    weekly_schedule: dict,
    special_days: dict,
) -> bytes:
    """
    Generate an FDTR .xlsx and return it as bytes.

    Parameters
    ----------
    weekly_schedule : {
        "monday": [{"time_in":"08:00","time_out":"12:00",
                    "category":"others","label":""},…], …
    }
    special_days : {
        "2026-01-01": {"type":"holiday",           "label":"NEW YEAR"},
        "2026-01-27": {"type":"leave",             "label":"ON SICK LEAVE"},
        "2026-01-20": {"type":"travel",            "label":"ON TRAVEL, TA NO: …"},
        "2026-02-06": {"type":"related_activities","label":""},
    }
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{MONTH_ABBR[month]} {year}"

    # ── Institution header (rows 1-9) ─────────────────────────────────────
    _write_institution_header(ws, month, year, faculty_name, department)

    # ── Column-header blocks ──────────────────────────────────────────────
    _write_col_headers(ws, COL_HEADER_ROW)
    _write_col_headers(ws, SECTION2_HEADER)

    # ── Day rows ──────────────────────────────────────────────────────────
    days_in_month = calendar.monthrange(year, month)[1]

    for day in range(1, days_in_month + 1):
        base_row  = _day_row(day)
        day_date  = date(year, month, day)
        date_str  = day_date.strftime("%Y-%m-%d")
        weekday   = day_date.weekday()   # 0=Mon … 6=Sun

        if date_str in special_days:
            entry    = special_days[date_str]
            day_type = entry["type"]
            label    = entry.get("label", "")

            if day_type == "related_activities":
                # Use custom times if provided, otherwise fall back to weekly schedule
                custom_in  = entry.get("time_in",  "").strip()
                custom_out = entry.get("time_out", "").strip()
                if custom_in and custom_out:
                    slots = [{"time_in": custom_in, "time_out": custom_out,
                              "category": "related_activities", "label": ""}]
                else:
                    weekday_name = WEEKDAY_NAMES[weekday]
                    slots = weekly_schedule.get(weekday_name, [])
                _write_regular_day(ws, base_row, day, slots,
                                   force_category="related_activities")
            elif day_type in ("leave", "travel"):
                _write_special_day(ws, base_row, day, label, 0)
            else:  # holiday
                _write_special_day(ws, base_row, day, label, None)

        elif weekday == 5:
            _write_special_day(ws, base_row, day, "SATURDAY", None)

        elif weekday == 6:
            _write_special_day(ws, base_row, day, "SUNDAY", None)

        else:
            weekday_name = WEEKDAY_NAMES[weekday]
            slots = weekly_schedule.get(weekday_name, [])
            _write_regular_day(ws, base_row, day, slots)

    # ── Footer ────────────────────────────────────────────────────────────
    _write_footer(ws, faculty_name, designation, dept_head, month, year)

    # ── Formatting ────────────────────────────────────────────────────────
    _set_column_widths(ws)
    _set_row_heights(ws, days_in_month)

    ws.freeze_panes = "A14"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = ws.PAPERSIZE_LEGAL
    ws.page_setup.fitToPage   = True
    ws.print_title_rows       = "1:13"

    # ── Serialize ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
